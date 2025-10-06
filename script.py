# Create a comprehensive Python solution for financial document analysis
# This will include all the functionality requested in the user query

import zipfile
import os
import pandas as pd
import pdfplumber
import re
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import streamlit as st
import shutil
from typing import List, Dict, Tuple, Optional
import tempfile

# Main Financial Document Analyzer Class
class FinancialDocumentAnalyzer:
    """
    A comprehensive tool for analyzing financial documents from ZIP files.
    Identifies missing schedules/annexures, blank pages, and validates financial totals.
    """
    
    def __init__(self):
        self.results = []
        self.missing_files = []
        self.required_schedules = [f"schedule {i}" for i in range(1, 23)]  # Schedule 1-22
        self.required_annexures = [f"annexure {i}" for i in range(1, 13)]  # Annexure 1-12
        
    def extract_zip_file(self, zip_path: str, extract_to: str = None) -> str:
        """Extract ZIP file and return extraction path"""
        if extract_to is None:
            extract_to = tempfile.mkdtemp()
            
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
        
        return extract_to
    
    def find_pdf_files(self, directory: str) -> List[str]:
        """Find all PDF files in directory and subdirectories"""
        pdf_files = []
        for root, dirs, files in os.walk(directory):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))
        return pdf_files
    
    def check_missing_files(self, pdf_files: List[str]) -> List[str]:
        """Check for missing schedule and annexure files"""
        found_files = [os.path.basename(f).lower() for f in pdf_files]
        missing = []
        
        # Check schedules
        for schedule in self.required_schedules:
            if not any(schedule in filename for filename in found_files):
                missing.append(schedule.title())
        
        # Check annexures  
        for annexure in self.required_annexures:
            if not any(annexure in filename for filename in found_files):
                missing.append(annexure.title())
        
        return missing
    
    def is_page_blank(self, text: str) -> bool:
        """Check if a page is essentially blank"""
        # Remove whitespace and check if meaningful content remains
        cleaned_text = re.sub(r'\s+', ' ', text.strip())
        return len(cleaned_text) < 50  # Threshold for blank page
    
    def extract_financial_tables(self, pdf_path: str) -> List[Dict]:
        """Extract tables and analyze financial data from PDF"""
        page_results = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text() or ""
                    
                    # Check if page is blank
                    is_blank = self.is_page_blank(page_text)
                    
                    # Extract tables
                    tables = page.extract_tables()
                    
                    # Analyze each table for financial columns
                    table_analysis = []
                    for table_idx, table in enumerate(tables):
                        if table and len(table) > 1:  # Must have headers and data
                            financial_data = self.analyze_financial_table(table)
                            if financial_data:
                                table_analysis.append({
                                    'table_index': table_idx + 1,
                                    'financial_data': financial_data
                                })
                    
                    page_results.append({
                        'page': page_num,
                        'is_blank': is_blank,
                        'tables': table_analysis,
                        'text_preview': page_text[:200] if page_text else ""
                    })
                    
        except Exception as e:
            print(f"Error processing {pdf_path}: {str(e)}")
            
        return page_results
    
    def analyze_financial_table(self, table: List[List]) -> Optional[Dict]:
        """Analyze table for financial columns and calculate totals"""
        if not table or len(table) < 2:
            return None
            
        headers = [str(cell).lower().strip() if cell else '' for cell in table[0]]
        
        # Look for financial column patterns
        financial_columns = {
            'opening_balance': -1,
            'debit': -1, 
            'credit': -1,
            'closing_balance': -1
        }
        
        # Match headers to financial columns
        for idx, header in enumerate(headers):
            if any(term in header for term in ['opening', 'opening balance', 'open bal']):
                financial_columns['opening_balance'] = idx
            elif any(term in header for term in ['debit', 'dr', 'debit amount']):
                financial_columns['debit'] = idx
            elif any(term in header for term in ['credit', 'cr', 'credit amount']):
                financial_columns['credit'] = idx
            elif any(term in header for term in ['closing', 'closing balance', 'close bal', 'balance']):
                financial_columns['closing_balance'] = idx
        
        # Only proceed if we found relevant financial columns
        if all(col == -1 for col in financial_columns.values()):
            return None
            
        # Calculate totals for each financial column
        totals = {}
        for col_name, col_idx in financial_columns.items():
            if col_idx >= 0:
                total = self.calculate_column_total(table, col_idx)
                totals[col_name] = total
        
        return {
            'column_mapping': financial_columns,
            'totals': totals,
            'row_count': len(table) - 1  # Excluding header
        }
    
    def calculate_column_total(self, table: List[List], col_idx: int) -> float:
        """Calculate total for a specific column"""
        total = 0.0
        for row in table[1:]:  # Skip header
            if col_idx < len(row) and row[col_idx]:
                cell_value = str(row[col_idx]).replace(',', '').replace('(', '-').replace(')', '')
                try:
                    # Extract numeric value
                    numeric_value = re.findall(r'-?\d+\.?\d*', cell_value)
                    if numeric_value:
                        total += float(numeric_value[0])
                except (ValueError, IndexError):
                    continue
        return total
    
    def check_receipt_payment_balance(self, page_results: List[Dict]) -> Dict:
        """Check if receipt and payment totals are equal on last page"""
        if not page_results:
            return {'status': 'No pages found', 'equal': False}
            
        last_page = page_results[-1]
        receipt_total = 0
        payment_total = 0
        
        for table in last_page.get('tables', []):
            financial_data = table.get('financial_data', {})
            totals = financial_data.get('totals', {})
            
            # Look for receipt/payment patterns
            if 'debit' in totals:
                receipt_total += totals['debit']
            if 'credit' in totals:
                payment_total += totals['credit']
        
        return {
            'receipt_total': receipt_total,
            'payment_total': payment_total,
            'equal': abs(receipt_total - payment_total) < 0.01,
            'difference': receipt_total - payment_total
        }
    
    def check_trial_balance_consistency(self, pdf_files: List[str]) -> Dict:
        """Check if trial balance pages have consistent grand totals"""
        trial_balance_files = [f for f in pdf_files if 'trial balance' in os.path.basename(f).lower()]
        
        if len(trial_balance_files) < 2:
            return {'status': 'Less than 2 trial balance files found', 'consistent': False}
            
        totals_comparison = []
        
        for tb_file in trial_balance_files[:2]:  # Check first two files
            page_results = self.extract_financial_tables(tb_file)
            grand_total = 0
            
            for page in page_results:
                for table in page.get('tables', []):
                    financial_data = table.get('financial_data', {})
                    totals = financial_data.get('totals', {})
                    grand_total += sum(totals.values())
            
            totals_comparison.append(grand_total)
        
        consistent = abs(totals_comparison[0] - totals_comparison[1]) < 0.01 if len(totals_comparison) == 2 else False
        
        return {
            'file1_total': totals_comparison[0] if len(totals_comparison) > 0 else 0,
            'file2_total': totals_comparison[1] if len(totals_comparison) > 1 else 0,
            'consistent': consistent,
            'difference': abs(totals_comparison[0] - totals_comparison[1]) if len(totals_comparison) == 2 else 0
        }
    
    def analyze_zip_file(self, zip_path: str) -> Dict:
        """Main analysis function"""
        # Extract ZIP file
        extract_path = self.extract_zip_file(zip_path)
        
        try:
            # Find PDF files
            pdf_files = self.find_pdf_files(extract_path)
            
            # Check for missing files
            missing_files = self.check_missing_files(pdf_files)
            
            # Analyze each PDF
            all_results = []
            for pdf_file in pdf_files:
                file_results = self.extract_financial_tables(pdf_file)
                
                # Compile results for this file
                file_summary = {
                    'filename': os.path.basename(pdf_file),
                    'pages': len(file_results),
                    'blank_pages': sum(1 for page in file_results if page['is_blank']),
                    'financial_tables': sum(len(page['tables']) for page in file_results),
                    'page_details': []
                }
                
                # Add page-level details
                for page in file_results:
                    page_detail = {
                        'page_number': page['page'],
                        'is_blank': page['is_blank'],
                        'opening_balance_total': 0,
                        'debit_total': 0,
                        'credit_total': 0,
                        'closing_balance_total': 0
                    }
                    
                    for table in page['tables']:
                        financial_data = table.get('financial_data', {})
                        totals = financial_data.get('totals', {})
                        
                        page_detail['opening_balance_total'] += totals.get('opening_balance', 0)
                        page_detail['debit_total'] += totals.get('debit', 0)
                        page_detail['credit_total'] += totals.get('credit', 0)
                        page_detail['closing_balance_total'] += totals.get('closing_balance', 0)
                    
                    file_summary['page_details'].append(page_detail)
                
                all_results.append(file_summary)
            
            # Check receipt/payment balance
            receipt_payment_check = {'status': 'No financial data found', 'equal': False}
            if pdf_files:
                # Use the first PDF for receipt/payment check
                first_pdf_results = self.extract_financial_tables(pdf_files[0])
                receipt_payment_check = self.check_receipt_payment_balance(first_pdf_results)
            
            # Check trial balance consistency
            trial_balance_check = self.check_trial_balance_consistency(pdf_files)
            
            return {
                'total_pdf_files': len(pdf_files),
                'missing_files': missing_files,
                'file_analysis': all_results,
                'receipt_payment_verification': receipt_payment_check,
                'trial_balance_verification': trial_balance_check
            }
            
        finally:
            # Clean up temporary directory
            shutil.rmtree(extract_path, ignore_errors=True)
    
    def generate_excel_report(self, analysis_results: Dict, output_path: str = "financial_analysis_report.xlsx"):
        """Generate Excel report with analysis results"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Summary sheet
        summary_ws = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_ws, analysis_results)
        
        # Create Detailed Analysis sheet
        detail_ws = wb.create_sheet("Detailed Analysis") 
        self._create_detailed_sheet(detail_ws, analysis_results)
        
        # Create Missing Files sheet
        missing_ws = wb.create_sheet("Missing Files")
        self._create_missing_files_sheet(missing_ws, analysis_results)
        
        # Create Verification sheet
        verification_ws = wb.create_sheet("Verification")
        self._create_verification_sheet(verification_ws, analysis_results)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, ws, results):
        """Create summary sheet in Excel"""
        # Headers
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        summary_data = [
            ["Total PDF Files", results['total_pdf_files']],
            ["Missing Files Count", len(results['missing_files'])],
            ["Receipt-Payment Balance", "Equal" if results['receipt_payment_verification']['equal'] else "Not Equal"],
            ["Trial Balance Consistency", "Consistent" if results['trial_balance_verification']['consistent'] else "Inconsistent"]
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2
    
    def _create_detailed_sheet(self, ws, results):
        """Create detailed analysis sheet"""
        headers = ["File Name", "Page", "Is Blank", "Opening Balance Total", 
                  "Debit Total", "Credit Total", "Closing Balance Total"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        row = 2
        for file_analysis in results['file_analysis']:
            filename = file_analysis['filename']
            for page_detail in file_analysis['page_details']:
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=page_detail['page_number'])
                ws.cell(row=row, column=3, value="Yes" if page_detail['is_blank'] else "No")
                ws.cell(row=row, column=4, value=round(page_detail['opening_balance_total'], 2))
                ws.cell(row=row, column=5, value=round(page_detail['debit_total'], 2))
                ws.cell(row=row, column=6, value=round(page_detail['credit_total'], 2))
                ws.cell(row=row, column=7, value=round(page_detail['closing_balance_total'], 2))
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def _create_missing_files_sheet(self, ws, results):
        """Create missing files sheet"""
        ws.cell(row=1, column=1, value="Missing Files").font = Font(bold=True)
        
        for row, missing_file in enumerate(results['missing_files'], 2):
            ws.cell(row=row, column=1, value=missing_file)
    
    def _create_verification_sheet(self, ws, results):
        """Create verification sheet"""
        headers = ["Verification Type", "Status", "Details"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Receipt-Payment verification
        rp_check = results['receipt_payment_verification']
        ws.cell(row=2, column=1, value="Receipt-Payment Balance")
        ws.cell(row=2, column=2, value="Equal" if rp_check['equal'] else "Not Equal")
        ws.cell(row=2, column=3, value=f"Receipt: {rp_check.get('receipt_total', 0)}, Payment: {rp_check.get('payment_total', 0)}")
        
        # Trial Balance verification
        tb_check = results['trial_balance_verification']
        ws.cell(row=3, column=1, value="Trial Balance Consistency")
        ws.cell(row=3, column=2, value="Consistent" if tb_check['consistent'] else "Inconsistent")
        ws.cell(row=3, column=3, value=f"File1 Total: {tb_check.get('file1_total', 0)}, File2 Total: {tb_check.get('file2_total', 0)}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

# Create a simple test to verify the class works
analyzer = FinancialDocumentAnalyzer()
print("✅ FinancialDocumentAnalyzer class created successfully!")
print(f"✅ Required schedules: {len(analyzer.required_schedules)} items")
print(f"✅ Required annexures: {len(analyzer.required_annexures)} items") 